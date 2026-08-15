#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Проверка Ведомости исполнения работ (ВИР) на несовместимые пары ТЦП 2621,
необоснованные коэффициенты и арифметические ошибки.

Использование:
    python3 check_vir.py <ведомость.xlsx> [--nds 0.22] [--verbose]

Источник правил: скилл tcp-2621 (references/tcp-2621-prices.md, tcp-2621-rules.md).

Проверки:
  1. Несовместимые пары пунктов (XOR — обе позиции в одной смете = двойная оплата).
  2. Запреты разделов (19.3.19 несовместим с разделами 3.7/3.18/3.19/18.1 и т.д.).
  3. Арифметика: Кол-во × Цена × Коэф = Стоимость.
  4. Сумма позиций = Итого.
  5. НДС = Итого × ставка (по умолчанию 22%).
  6. Коэффициенты на сверку с известными значениями ТЦП.
"""

import argparse
import re
import sys
from pathlib import Path

# --- Несовместимые пары ТЦП (пункт ↔ пункт) ---
INCOMPAT_PAIRS = {
    ("1.31", "1.35"),
    ("1.37", "1.40"),
    ("2.21", "2.24"),
    ("2.22", "2.25"),
    ("2.51", "2.48"),
    ("2.51", "2.49"),
    ("3.1.13", "3.1.1"),
    ("3.16", "3.10"),
    ("20.1", "1.39"),
}

# --- Запреты разделов: пункт → префиксы разделов, с которыми несовместим ---
# Префикс '3.7.' означает все пункты 3.7.x; '18.1.' — все пункты 18.1.x.
SECTION_BANS = {
    "19.3.19": ["3.7.", "3.18.", "3.19.", "18.1.", "3.6."],
    "17.1.3": ["3.7."],
}

# --- Коэффициенты для сверки: пункт → (ожидаемое значение, условие) ---
KNOWN_COEFFS = {
    "1.33.1": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.2": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.3": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.5": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.6": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.7": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.8": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.9": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.10": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.12": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.13": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.33.15": (0.75, "перевыпуск в электронном редактируемом формате"),
    "1.46": (0.8, "без регистрации в Ростехнадзоре (по согласованию)"),
    "1.47": (0.8, "без регистрации в Ростехнадзоре (по согласованию)"),
    "2.28": (0.75, "каждый последующий термобокс (к первому)"),
    "2.40.1": (0.6, "только поставка (без монтажа)"),
    "2.40.2": (0.6, "только поставка (без монтажа)"),
    "2.59": (0.4, "демонтаж/монтаж существующих кожухов при модернизации"),
    "4.13": (0.5, "ПНР ЭПУ/ИБП при изменении ёмкости АКБ"),
    "5.8.3": (1.2, "прокладка в стальной трубе"),
    "5.8.4": (1.2, "прокладка в стальной трубе"),
    "5.8.5": (1.2, "прокладка в стальной трубе"),
    "5.8.6": (1.2, "прокладка в стальной трубе"),
    "5.25": (1.2, "прокладка в стальной трубе"),
    "9.1": (0.75, "демонтаж (с доставкой на склад по заданию Заказчика)"),
    "9.1.1": (1.5, "демонтаж/монтаж (перенос) к пунктам монтажа"),
    "9.1.2": (1.75, "демонтаж/монтаж (перенос) с доставкой на новое место"),
    "9.1.3": (0.01, "демонтаж ВЧ-фидера без сдачи на склад (к монтажу фидера р.3.6)"),
    "9.4": (0.75, "% от п.п. 2.22, 2.25"),
    "9.4.2": (1.75, "% от п.п. 2.22, 2.25"),
    "9.8": (0.75, "% от стоимости монтажа мет. конструкций"),
    "9.8.1": (1.25, "% от стоимости монтажа мет. конструкций"),
    "22.3.5": (1.2, "автовышка высокой проходимости"),
    "22.3.6": (1.2, "автовышка высокой проходимости"),
    "22.3.7": (1.2, "автовышка высокой проходимости"),
    "22.3.8": (1.2, "автовышка высокой проходимости"),
    "22.3.10": (1.2, "автовышка высокой проходимости"),
}

# --- Коэффициенты, требующие проверки условия (не жёсткое ожидание) ---
# Пункт → (значение при условии, описание условия)
CONDITIONAL_COEFFS = {
    "3.10": (1.5, "сдвоенный; для одинарного — 1,0"),
    "3.11": (1.5, "сдвоенный комбайнер; для одинарного — 1,0"),
    "3.12": (1.5, "сдвоенный МШУ; для одинарного — 1,0"),
    "3.1.9": (1.0, "монтаж; при демонтаже применять 9.1 = 0,75"),
    "3.7.4": (0.75, "монтаж АФУ с подключением к существующим антеннам (без новых антенн); если антенны новые — 1,0"),
}

NUM_RE = re.compile(r'^\d+(\.\d+)*$')


def norm_nums(s):
    """Разбить '3.7.4/ 9.1' → ['3.7.4', '9.1']; '1.33.3' → ['1.33.3']."""
    if not s:
        return []
    s = s.strip()
    parts = [p.strip() for p in re.split(r'[/,;]+', s) if p.strip()]
    out = []
    for p in parts:
        # убрать хвосты вида «9.1.1» из «19.3.19/ 9.1.1»
        if NUM_RE.match(p):
            out.append(p)
    return out


def parse_vir(path):
    """Распарсить ВИР: листы, строки с номером ТЦП, кол-вом, ценой, коэф., стоимостью."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    positions = []
    totals = {"sum": None, "nds": None}

    for ws in wb.worksheets:
        # Ищем строку-заголовок: содержит «ТЦП» и «Наименование»
        header_row = None
        for r in range(1, min(ws.max_row, 30) + 1):
            vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
            if any('ТЦП' in v for v in vals) and any('Наименование' in v for v in vals):
                header_row = r
                break
        if header_row is None:
            continue

        # Маппинг колонок по заголовку
        cols = {}
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(header_row, c).value or '').strip()
            if 'ТЦП' in v and ('№' in v or 'п.' in v or 'п/п' in v):
                cols['num'] = c
            elif 'Наименование' in v:
                cols['name'] = c
            elif ('Ед' in v and 'изм' in v) or v == 'Ед. изм.':
                cols['unit'] = c
            elif 'Кол-во' in v or v == 'Кол.' or v == 'Кол-во, шт':
                cols['qty'] = c
            elif 'Цена за единицу' in v or ('Цена' in v and 'единиц' in v):
                cols['price'] = c
            elif 'Коэф' in v:
                cols['coef'] = c
            elif 'Полная стоимость' in v or ('Стоимость' in v and 'полн' in v.lower()):
                cols['total'] = c

        def cell(r, key):
            c = cols.get(key)
            if c is None:
                return None
            return ws.cell(r, c).value

        for r in range(header_row + 1, ws.max_row + 1):
            num_raw = str(cell(r, 'num') or '').strip()
            name = str(cell(r, 'name') or '').strip()

            # Строка итога — только точные маркеры (иначе «аналогичную» в названии даёт ложное срабатывание)
            text = (num_raw + ' ' + name).lower()
            if ('итого' in text and 'ндс' not in text) or 'итого без ндс' in text:
                totals['sum'] = cell(r, 'total')
                continue
            if 'налог на добавленную' in text or text.strip().startswith('ндс') or 'итого с ндс' in text:
                totals['nds'] = cell(r, 'total')
                continue

            # Позиция: нужен номер ТЦП (цифровой) или «X/ Y» с номером
            nums = norm_nums(num_raw)
            if not nums and not name:
                continue
            # Заголовки типа «Работы на БС: …» без номера ТЦП — пропускаем
            if not nums:
                continue

            def fval(key):
                v = cell(r, key)
                if v is None or v == '':
                    return None
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return None

            positions.append({
                "row": r,
                "sheet": ws.title,
                "num": num_raw,
                "nums": nums,
                "name": name[:120],
                "qty": fval('qty') if 'qty' in cols else 1.0,
                "price": fval('price'),
                "coef": fval('coef') if fval('coef') is not None else 1.0,
                "total": fval('total'),
            })
    return positions, totals


def check_incompat(positions):
    findings = []
    # Пары пункт-пункт
    for i in range(len(positions)):
        for j in range(i + 1, len(positions)):
            a, b = positions[i], positions[j]
            for na in a["nums"]:
                for nb in b["nums"]:
                    if (na, nb) in INCOMPAT_PAIRS or (nb, na) in INCOMPAT_PAIRS:
                        findings.append({
                            "type": "Несовместимая пара",
                            "detail": f"{na} ↔ {nb}",
                            "amount": None,
                            "name": f"стр. {a['row']}: {a['name'][:50]} | стр. {b['row']}: {b['name'][:50]}",
                        })
    # Запреты разделов (по префиксам)
    for p in positions:
        for np_ in p["nums"]:
            if np_ not in SECTION_BANS:
                continue
            for prefix in SECTION_BANS[np_]:
                for q in positions:
                    if q is p:
                        continue
                    for nq in q["nums"]:
                        if nq.startswith(prefix):
                            findings.append({
                                "type": "Запрет раздела",
                                "detail": f"{np_} несовместим с разделом {prefix[:-1]} (позиция {nq})",
                                "amount": None,
                                "name": f"стр. {p['row']}: {p['name'][:50]} | стр. {q['row']}: {q['name'][:50]}",
                            })
    return findings


def main():
    ap = argparse.ArgumentParser(description="Проверка ВИР на несовместимости ТЦП, коэффициенты и арифметику")
    ap.add_argument("vir", help="Путь к xlsx ведомости")
    ap.add_argument("--nds", type=float, default=0.22, help="Ставка НДС (по умолчанию 0.22)")
    ap.add_argument("--verbose", action="store_true", help="Показать все позиции")
    args = ap.parse_args()

    if not Path(args.vir).exists():
        print(f"Файл не найден: {args.vir}")
        sys.exit(1)

    positions, totals = parse_vir(args.vir)
    if not positions:
        print("Позиции не найдены. Проверьте, что в файле есть строка с «№ п. по ТЦП» и «Наименование».")
        sys.exit(1)

    print(f"=== ВИР: {args.vir} ===")
    print(f"Позиций: {len(positions)}\n")

    # 1. Несовместимости
    findings = check_incompat(positions)
    if findings:
        print("!!! НЕСОВМЕСТИМОСТИ:")
        for f in findings:
            print(f"  [{f['type']}] {f['detail']}")
            print(f"      {f['name']}")
        print()
    else:
        print("Несовместимых пар: нет\n")

    # 2. Арифметика
    print("--- Арифметика позиций ---")
    sum_total = 0.0
    arith_errors = 0
    for p in positions:
        if p["price"] is None:
            continue
        calc = round(p["qty"] * p["price"] * p["coef"], 2)
        sum_total += calc
        flag = ""
        if p["total"] is not None and abs(calc - p["total"]) > 0.01:
            flag = f"  <-- ОШИБКА: в ведомости {p['total']}"
            arith_errors += 1
        if args.verbose:
            print(f"  [стр.{p['row']}] {p['num']:<10} {p['name'][:48]} | {p['qty']} × {p['price']} × {p['coef']} = {calc}{flag}")
    if not args.verbose:
        print(f"  (детализация опущена; ошибок арифметики: {arith_errors})")
    print()

    # 3. Итог и НДС
    print("--- Итог ---")
    print(f"  Сумма позиций:      {sum_total:,.2f}")
    decl_sum = totals.get('sum')
    if decl_sum is not None:
        try:
            diff = round(float(decl_sum) - sum_total, 2)
            status = "OK" if abs(diff) < 0.01 else f"РАСХОЖДЕНИЕ {diff:+.2f}"
            print(f"  Итого в ведомости:  {float(decl_sum):,.2f}  [{status}]")
        except (TypeError, ValueError):
            pass
    decl_nds = totals.get('nds')
    if decl_nds is not None:
        try:
            expect_nds = round(sum_total * args.nds, 2)
            diff = round(float(decl_nds) - expect_nds, 2)
            status = "OK" if abs(diff) < 0.01 else f"РАСХОЖДЕНИЕ {diff:+.2f}"
            print(f"  НДС в ведомости:    {float(decl_nds):,.2f} (ожидалось {expect_nds:,.2f} при {args.nds:.0%}) [{status}]")
        except (TypeError, ValueError):
            pass

    # 4. Коэффициенты
    print("\n--- Коэффициенты (сверка с ТЦП) ---")
    issues = 0
    for p in positions:
        # «X/9.1», «X/9.1.1» — демонтажные модификаторы: коэффициент = произведение
        # (напр. 0,75 демонтажа × 0,75 скидки монтажа АФУ = 0,5625 — корректно).
        has_demont = any(n.startswith('9.') for n in p["nums"])
        for n in p["nums"]:
            if has_demont and n not in KNOWN_COEFFS:
                # базовый пункт в демонтажной позиции — его коэффициент умножается на демонтажный
                continue
            if n in KNOWN_COEFFS:
                expected, cond = KNOWN_COEFFS[n]
                if has_demont and n.startswith('9.'):
                    # демонтажный модификатор: итоговый коэф = произведение (напр. 0,75 × 0,75 = 0,5625)
                    base = [m for m in p["nums"] if not m.startswith('9.')]
                    if base:
                        base_coef = (CONDITIONAL_COEFFS.get(base[0], KNOWN_COEFFS.get(base[0], (1, '')))[0])
                        expected_prod = round(base_coef * expected, 4)
                        if abs(p["coef"] - expected_prod) > 0.001:
                            print(f"  [стр.{p['row']}] {n}: коэф {p['coef']} ≠ {expected_prod} = {base[0]}×{expected} ({cond})")
                            issues += 1
                        continue
                if abs(p["coef"] - expected) > 0.001:
                    print(f"  [стр.{p['row']}] {n}: коэф {p['coef']} ≠ ожидаемый {expected} ({cond})")
                    issues += 1
            if n in CONDITIONAL_COEFFS:
                val_when, cond = CONDITIONAL_COEFFS[n]
                if abs(p["coef"] - val_when) < 0.001:
                    print(f"  [стр.{p['row']}] {n}: коэф {p['coef']} — проверить: {cond}")
                else:
                    print(f"  [стр.{p['row']}] {n}: коэф {p['coef']} ≠ {val_when} — проверить условие: {cond}")
                    issues += 1
    if issues == 0:
        print("  Все проверенные коэффициенты соответствуют ТЦП.")

    sys.exit(0 if (not findings and arith_errors == 0 and issues == 0) else 2)


if __name__ == "__main__":
    main()
